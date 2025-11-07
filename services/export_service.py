# services/export_service.py
"""
Export Service voor Planning Tool
Genereert Excel exports van planning voor HR
v0.6.20: Validatie Rapport tab toegevoegd
"""

from pathlib import Path
from datetime import datetime, date
from database.connection import get_connection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from services.bemannings_controle_service import controleer_maand, format_ontbrekende_codes, format_dubbele_codes
import calendar


# Nederlandse maandnamen
MAAND_NAMEN = {
    1: 'januari', 2: 'februari', 3: 'maart', 4: 'april',
    5: 'mei', 6: 'juni', 7: 'juli', 8: 'augustus',
    9: 'september', 10: 'oktober', 11: 'november', 12: 'december'
}


def get_planner_notities_voor_maand(jaar: int, maand: int) -> dict:
    """
    Haal planner notities op voor een maand

    Returns:
        Dict met datum_str als key en lijst van notities als value
        {'2025-01-15': ['Jan: Ruil met Piet', 'Marie: Opleiding']}
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Haal alle notities op voor deze maand (notities beginnen met "[Planner]:")
    cursor.execute("""
        SELECT p.datum, g.volledige_naam, p.notitie
        FROM planning p
        JOIN gebruikers g ON p.gebruiker_id = g.id
        WHERE strftime('%Y', p.datum) = ?
        AND strftime('%m', p.datum) = ?
        AND p.notitie IS NOT NULL
        AND p.notitie != ''
        AND p.notitie LIKE '[Planner]:%'
        ORDER BY p.datum, g.volledige_naam
    """, (str(jaar), str(maand).zfill(2)))

    notities_per_dag = {}
    for row in cursor.fetchall():
        datum_str = row['datum']
        naam = row['volledige_naam']
        notitie = row['notitie']

        # Verwijder "[Planner]: " prefix voor display
        notitie_text = notitie.replace('[Planner]: ', '').strip()

        if datum_str not in notities_per_dag:
            notities_per_dag[datum_str] = []

        notities_per_dag[datum_str].append(f"{naam}: {notitie_text}")

    conn.close()
    return notities_per_dag


def export_maand_naar_excel(jaar: int, maand: int) -> str:
    """
    Exporteer planning van een maand naar Excel bestand

    Args:
        jaar: Jaar (bijv. 2025)
        maand: Maand nummer (1-12)

    Returns:
        str: Pad naar gegenereerd bestand
    """
    # Exports directory (wordt aangemaakt bij app start)
    export_dir = Path("exports")

    # Bestandsnaam: maandnaam_jaartal.xlsx
    maand_naam = MAAND_NAMEN[maand]
    bestand_naam = f"{maand_naam}_{jaar}.xlsx"
    bestand_pad = export_dir / bestand_naam

    # Haal planning data op
    planning_data = haal_planning_data(jaar, maand)

    # Maak Excel bestand
    wb = Workbook()
    ws = wb.active
    ws.title = f"{maand_naam.capitalize()} {jaar}"

    # Styling definities
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Bepaal alle dagen in de maand
    dagen_in_maand = calendar.monthrange(jaar, maand)[1]
    datums = []
    for dag in range(1, dagen_in_maand + 1):
        datum = datetime(jaar, maand, dag)
        datums.append(datum)

    # Header rij 1: Titel
    ws.merge_cells('A1:B1')
    titel_cell = ws['A1']
    titel_cell.value = f"Planning {maand_naam.capitalize()} {jaar}"
    titel_cell.font = Font(bold=True, size=14, color="366092")
    titel_cell.alignment = center_alignment

    # Header rij 2: Kolom headers
    ws['A2'] = "Naam"
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].border = border
    ws['A2'].alignment = center_alignment

    # Datum kolommen
    for col_idx, datum in enumerate(datums, start=2):
        cell = ws.cell(row=2, column=col_idx)
        # Formaat: "Ma 1" of "Di 2" etc.
        dag_naam = ['Ma', 'Di', 'Wo', 'Do', 'Vr', 'Za', 'Zo'][datum.weekday()]
        cell.value = f"{dag_naam}\n{datum.day}"
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data rijen
    row_idx = 3
    for gebruiker_data in planning_data:
        # Naam kolom
        naam_cell = ws.cell(row=row_idx, column=1)
        naam_cell.value = gebruiker_data['naam']
        naam_cell.font = Font(size=10)
        naam_cell.border = border
        naam_cell.alignment = Alignment(vertical='center')

        # Shift codes per dag
        for col_idx, datum in enumerate(datums, start=2):
            datum_str = datum.strftime('%Y-%m-%d')
            shift_code = gebruiker_data['planning'].get(datum_str, '')

            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = shift_code
            cell.font = Font(size=10)
            cell.border = border
            cell.alignment = center_alignment

            # Achtergrondkleur voor weekend/feestdag
            if datum.weekday() >= 5:  # Zaterdag of Zondag
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        row_idx += 1

    # Kolom breedtes aanpassen
    ws.column_dimensions['A'].width = 25  # Naam kolom
    for col_idx in range(2, len(datums) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8  # Datum kolommen

    # Rij hoogtes
    ws.row_dimensions[1].height = 25  # Titel
    ws.row_dimensions[2].height = 30  # Header

    # Voeg Validatie Rapport sheet toe (v0.6.20)
    validatie_ws = wb.create_sheet(title="Validatie Rapport")
    maak_validatie_rapport_sheet(validatie_ws, jaar, maand)

    # Sla op (v0.6.21: verbeterde error handling)
    try:
        wb.save(bestand_pad)
    except PermissionError:
        # Re-raise met bestandsnaam voor duidelijkere error message
        raise PermissionError(f"Kan bestand niet opslaan: {bestand_naam}. Bestand is waarschijnlijk open in Excel.")
    except (IOError, OSError) as e:
        # Re-raise met context
        raise IOError(f"Kan bestand niet schrijven naar {bestand_pad}: {str(e)}")

    return str(bestand_pad)


def haal_planning_data(jaar: int, maand: int) -> list:
    """
    Haal planning data op voor een maand
    Inclusief reserves (is_reserve=1)

    Returns:
        List van dicts met naam en planning per datum
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Haal alle gebruikers op inclusief reserves, exclusief admin
    cursor.execute("""
        SELECT id, volledige_naam, is_reserve
        FROM gebruikers
        WHERE is_actief = 1
        AND gebruikersnaam != 'admin'
        ORDER BY is_reserve ASC, volledige_naam ASC
    """)

    gebruikers = cursor.fetchall()

    # Bepaal datum range
    eerste_dag = f"{jaar:04d}-{maand:02d}-01"
    if maand == 12:
        volgende_maand = f"{jaar + 1:04d}-01-01"
    else:
        volgende_maand = f"{jaar:04d}-{maand + 1:02d}-01"

    planning_data = []

    for gebruiker in gebruikers:
        gebruiker_id = gebruiker['id']
        naam = gebruiker['volledige_naam']

        # Haal planning op voor deze gebruiker
        cursor.execute("""
            SELECT datum, shift_code
            FROM planning
            WHERE gebruiker_id = ?
            AND datum >= ?
            AND datum < ?
            ORDER BY datum
        """, (gebruiker_id, eerste_dag, volgende_maand))

        planning_records = cursor.fetchall()

        # Maak dict van datum -> shift_code
        planning_dict = {}
        for record in planning_records:
            planning_dict[record['datum']] = record['shift_code']

        planning_data.append({
            'naam': naam,
            'planning': planning_dict
        })

    conn.close()

    return planning_data


def maak_validatie_rapport_sheet(ws, jaar: int, maand: int) -> None:
    """
    Vul validatie rapport sheet met bemannings controle resultaten.

    Args:
        ws: Openpyxl worksheet object
        jaar: Jaar (bijv. 2025)
        maand: Maand nummer (1-12)
    """
    # Styling definities
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    groen_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green
    oranje_fill = PatternFill(start_color="FFB74D", end_color="FFB74D", fill_type="solid")  # Intenser orange (Material Orange 300)
    rood_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Light red
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    maand_naam = MAAND_NAMEN[maand]

    # Titel rij 1
    ws.merge_cells('A1:E1')
    titel_cell = ws['A1']
    titel_cell.value = f"Bemannings Validatie (Kritische Shifts) - {maand_naam.capitalize()} {jaar}"
    titel_cell.font = Font(bold=True, size=14, color="366092")
    titel_cell.alignment = center_alignment

    # Header rij 2
    headers = ["Datum", "Dag", "Status", "Ontbrekende Kritische Shifts", "Planner Notities"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Haal bemannings controle resultaten op
    validatie_resultaat = controleer_maand(jaar, maand)
    dagen_resultaten = validatie_resultaat['dagen']
    samenvatting = validatie_resultaat['samenvatting']

    # Haal planner notities op
    planner_notities = get_planner_notities_voor_maand(jaar, maand)

    # Data rijen
    row_idx = 3
    dagen_in_maand = calendar.monthrange(jaar, maand)[1]

    for dag in range(1, dagen_in_maand + 1):
        datum = date(jaar, maand, dag)
        datum_str = datum.strftime('%Y-%m-%d')

        # Haal resultaat op voor deze dag
        if datum_str in dagen_resultaten:
            dag_resultaat = dagen_resultaten[datum_str]
            status = dag_resultaat['status']
            ontbrekende = dag_resultaat.get('ontbrekende_codes', [])
            dubbele = dag_resultaat.get('dubbele_codes', [])

            # Check of er notities zijn voor deze dag
            heeft_notities = datum_str in planner_notities

            # Toon alleen dagen met ontbrekende shifts OF planner notities
            if status != 'rood' and not heeft_notities:
                continue

            # Datum kolom
            datum_cell = ws.cell(row=row_idx, column=1)
            datum_cell.value = datum.strftime('%d-%m-%Y')
            datum_cell.border = border
            datum_cell.alignment = center_alignment

            # Dag kolom
            dag_naam = ['Ma', 'Di', 'Wo', 'Do', 'Vr', 'Za', 'Zo'][datum.weekday()]
            dag_cell = ws.cell(row=row_idx, column=2)
            dag_cell.value = dag_naam
            dag_cell.border = border
            dag_cell.alignment = center_alignment

            # Status kolom
            status_cell = ws.cell(row=row_idx, column=3)
            if status == 'groen':
                status_cell.value = "✓ Volledig"
                status_cell.fill = groen_fill
            elif status == 'geel':
                status_cell.value = "⚠ Dubbel"
                status_cell.fill = oranje_fill
            else:  # rood
                status_cell.value = "✗ Onvolledig"
                status_cell.fill = rood_fill
            status_cell.border = border
            status_cell.alignment = center_alignment

            # Ontbrekende shifts kolom
            ontbrekend_cell = ws.cell(row=row_idx, column=4)
            if ontbrekende:
                ontbrekend_cell.value = format_ontbrekende_codes(ontbrekende)
            elif dubbele:
                ontbrekend_cell.value = format_dubbele_codes(dubbele)
            else:
                ontbrekend_cell.value = ""
            ontbrekend_cell.border = border
            ontbrekend_cell.alignment = left_alignment

            # Planner notities kolom
            notities_cell = ws.cell(row=row_idx, column=5)
            if heeft_notities:
                notities_text = "\n".join(planner_notities[datum_str])
                notities_cell.value = notities_text
            else:
                notities_cell.value = ""
            notities_cell.border = border
            notities_cell.alignment = left_alignment

            row_idx += 1

    # Samenvatting sectie (na data rijen + 2 lege rijen)
    samenvatting_start_row = row_idx + 2

    # Streep
    ws.merge_cells(f'A{samenvatting_start_row}:E{samenvatting_start_row}')
    streep_cell = ws.cell(row=samenvatting_start_row, column=1)
    streep_cell.value = "─" * 60
    streep_cell.font = Font(size=10)

    # Titel samenvatting
    titel_row = samenvatting_start_row + 1
    ws.merge_cells(f'A{titel_row}:E{titel_row}')
    titel_cell = ws.cell(row=titel_row, column=1)
    titel_cell.value = "TOTAAL OVERZICHT"
    titel_cell.font = Font(bold=True, size=12)

    # Streep
    streep2_row = titel_row + 1
    ws.merge_cells(f'A{streep2_row}:E{streep2_row}')
    streep2_cell = ws.cell(row=streep2_row, column=1)
    streep2_cell.value = "─" * 60
    streep2_cell.font = Font(size=10)

    # Samenvatting cijfers
    cijfers_start = streep2_row + 1

    samenvatting_items = [
        (f"Volledig bemand:", f"{samenvatting['volledig']} dagen", groen_fill),
        (f"Dubbele codes:", f"{samenvatting['dubbel']} dagen", oranje_fill),
        (f"Onvolledig bemand:", f"{samenvatting['onvolledig']} dagen", rood_fill),
        (f"TOTAAL:", f"{samenvatting['totaal']} dagen", None)
    ]

    for idx, (label, waarde, fill) in enumerate(samenvatting_items):
        row = cijfers_start + idx

        label_cell = ws.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)

        waarde_cell = ws.cell(row=row, column=2)
        waarde_cell.value = waarde
        waarde_cell.font = Font(size=10)

        if fill:
            label_cell.fill = fill
            waarde_cell.fill = fill

    # Kolom breedtes
    ws.column_dimensions['A'].width = 15  # Datum
    ws.column_dimensions['B'].width = 8   # Dag
    ws.column_dimensions['C'].width = 15  # Status
    ws.column_dimensions['D'].width = 50  # Ontbrekende kritische shifts
    ws.column_dimensions['E'].width = 50  # Planner notities

    # Rij hoogtes
    ws.row_dimensions[1].height = 25  # Titel
    ws.row_dimensions[2].height = 30  # Header
